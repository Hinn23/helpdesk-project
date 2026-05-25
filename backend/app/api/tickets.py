import csv, io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from sqlalchemy.orm import Session
from app.sse import events as sse_events

from app.database import get_db
from app.schemas.ticket import TicketCreate, TicketUpdate, TicketRead
from app.services.ticket_service import TicketService
from app.services.comment_service import CommentService
from app.auth import get_optional_user, get_current_user
from app.models.user import User
from app.models.category import Category
from app.models.subscription import Subscription
from app.models.audit_log import AuditLog

router = APIRouter(prefix="/tickets", tags=["tickets"])


@router.get("/", response_model=dict)
def list_tickets(
    skip: int = Query(0, ge=0),
    limit: int = Query(10, ge=1, le=100),
    status: Optional[str] = None,
    priority: Optional[str] = None,
    search: Optional[str] = None,
    category_id: Optional[int] = None,
    assignee_id: Optional[int] = None,
    author_id: Optional[int] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    sort: Optional[str] = "created_at",
    order: Optional[str] = "desc",
    db: Session = Depends(get_db),
    current_user=Depends(get_optional_user),
):
    svc = TicketService(db)
    is_admin = current_user and current_user.role in ("admin", "manager")
    if not is_admin:
        if status == "on_moderation":
            if current_user:
                items_raw, total = svc.get_all_filtered(
                    skip=skip, limit=limit, status="on_moderation",
                    author_id=current_user.id,
                )
            else:
                return {"items": [], "total": 0, "page": 1, "limit": limit, "pages": 0}
        else:
            items_raw_all, total = svc.get_all_filtered(
                skip=0, limit=100000, status=status, priority=priority,
                search=search, category_id=category_id, assignee_id=assignee_id,
                author_id=author_id, date_from=date_from, date_to=date_to,
                sort=sort, order=order,
            )
            user_id = current_user.id if current_user else -1
            items_raw = [t for t in items_raw_all if t.status != "on_moderation" or t.author_id == user_id]
            total = len(items_raw)
            items_raw = items_raw[skip:skip + limit]
    else:
        items_raw, total = svc.get_all_filtered(
            skip=skip, limit=limit, status=status, priority=priority,
            search=search, category_id=category_id, assignee_id=assignee_id,
            author_id=author_id, date_from=date_from, date_to=date_to,
            sort=sort, order=order,
        )
    return {
        "items": [TicketRead.model_validate(t) for t in items_raw],
        "total": total,
        "page": skip // limit + 1,
        "limit": limit,
        "pages": (total + limit - 1) // limit if total else 0,
    }


@router.get("/{ticket_id}", response_model=TicketRead)
def get_ticket(ticket_id: int, db: Session = Depends(get_db), current_user=Depends(get_optional_user)):
    try:
        ticket = TicketService(db).get_by_id(ticket_id)
        if ticket.status == "on_moderation" and not (current_user and current_user.role in ("admin", "manager")):
            if not current_user or ticket.author_id != current_user.id:
                raise HTTPException(404, f"Заявка #{ticket_id} не найдена")
        return ticket
    except ValueError:
        raise HTTPException(404, f"Заявка #{ticket_id} не найдена. Может, её удалили? Или не было вовсе?")


@router.post("/", response_model=TicketRead, status_code=201)
def create_ticket(data: TicketCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role == "guest":
        raise HTTPException(403, "Гости не могут создавать заявки. Только смотреть подглядывать")
    update_kwargs = {"author_id": current_user.id}
    if current_user.role not in ("admin", "manager"):
        update_kwargs["status"] = "on_moderation"
        update_kwargs["priority"] = "medium"
    ticket_data = data.model_copy(update=update_kwargs)
    try:
        result = TicketService(db).create(ticket_data, user_name=current_user.name)
        sse_events.broadcast_sync("ticket_created", {"id": result.id, "title": result.title})
        return result
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.put("/{ticket_id}", response_model=TicketRead)
def update_ticket(ticket_id: int, data: TicketUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    svc = TicketService(db)
    try:
        ticket = svc.get_by_id(ticket_id)
    except ValueError:
        raise HTTPException(404, f"Заявка #{ticket_id} не найдена. Куда ты тыкаешь?")
    if current_user.role == "guest":
        raise HTTPException(403, "Гости не могут редактировать заявки")
    if ticket.status == "on_moderation" and current_user.role not in ("admin", "manager"):
        raise HTTPException(403, "Заявка на модерации. Дождитесь проверки администратором.")
    if ticket.status in ("done", "closed"):
        label = {"done": "завершена", "closed": "закрыта"}.get(ticket.status, ticket.status)
        raise HTTPException(403, f"Заявка {label}. Только персонал может вносить изменения.")
    if current_user.role not in ("admin", "manager", "support") and ticket.author_id != current_user.id:
        raise HTTPException(403, "Только админ, менеджер, саппорт или автор может редактировать заявку")
    if current_user.role == "support" and ticket.author_id != current_user.id:
        allowed = {"status", "priority", "assignee_id"}
        updating = set(data.model_dump(exclude_unset=True).keys())
        if not updating.issubset(allowed):
            raise HTTPException(403, "Саппорт может менять только статус, приоритет и исполнителя")
    result = svc.update(ticket_id, data, user_name=current_user.name)
    changed_fields = [k for k in data.model_dump(exclude_unset=True) if k != 'author_id']
    if changed_fields:
        sse_events.broadcast_sync("ticket_updated", {"id": result.id, "title": result.title, "changed": changed_fields})

    if changed_fields:
        subs = db.query(Subscription).filter(Subscription.ticket_id == ticket_id, Subscription.user_id != current_user.id).all()
        for s in subs:
            db.add(AuditLog(
                ticket_id=ticket_id, user_name=current_user.name,
                action="notified", field="subscription",
                new_value=f"Изменено: {', '.join(changed_fields)}",
            ))
        db.commit()

    return result


@router.delete("/{ticket_id}", status_code=204)
def delete_ticket(ticket_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    svc = TicketService(db)
    try:
        ticket = svc.get_by_id(ticket_id)
    except ValueError:
        raise HTTPException(404, f"Заявка #{ticket_id} не найдена. Нечего удалять")
    if current_user.role not in ("admin", "manager") and ticket.author_id != current_user.id:
        raise HTTPException(403, "Только админ, менеджер или автор может удалять заявку")
    if current_user.role == "support":
        raise HTTPException(403, "Саппорт не может удалять заявки")
    svc.delete(ticket_id)


@router.post("/batch-delete", status_code=200)
def batch_delete(data: dict, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(403, "Только админ может массово удалять заявки")
    ids = data.get("ids", [])
    svc = TicketService(db)
    deleted = 0
    for tid in ids:
        try:
            svc.get_by_id(tid)
            svc.delete(tid)
            deleted += 1
        except ValueError:
            continue
    return {"deleted": deleted}


@router.post("/batch-status", status_code=200)
def batch_status(data: dict, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(403, "Только админ может массово менять статус")
    ids = data.get("ids", [])
    new_status = data.get("status", "")
    allowed = {"new", "in_progress", "done", "cancelled", "on_moderation"}
    if new_status not in allowed:
        raise HTTPException(400, f"Недопустимый статус: {new_status}")
    svc = TicketService(db)
    updated = 0
    for tid in ids:
        try:
            ticket = svc.get_by_id(tid)
            update_data = {"status": new_status}
            import types
            update_obj = types.SimpleNamespace(**{"model_dump": lambda exclude_unset=True: update_data})
            svc.update(tid, update_obj, user_name=current_user.name)
            updated += 1
        except ValueError:
            continue
    return {"updated": updated}


@router.get("/export/csv")
def export_csv(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    category_id: Optional[int] = None,
    assignee_id: Optional[int] = None,
    author_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(get_optional_user),
):
    svc = TicketService(db)
    comment_svc = CommentService(db)
    items_raw, _ = svc.get_all_filtered(skip=0, limit=10000, status=status, priority=priority, category_id=category_id, assignee_id=assignee_id, author_id=author_id)

    users_map = {u.id: u.name for u in db.query(User).all()}
    cats_map = {c.id: c.name for c in db.query(Category).all()}

    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "ID", "Название", "Описание", "Статус", "Приоритет",
        "Дедлайн", "Автор", "Исполнитель", "Категория", "Создана", "Обновлена",
        "Комментарии",
    ])
    for t in items_raw:
        comments = comment_svc.get_by_ticket(t.id)
        comments_text = " | ".join(
            f"{c.author_name} ({c.created_at.strftime('%d.%m.%Y')}): {c.text}"
            for c in comments
        ) if comments else ""
        writer.writerow([
            t.id, t.title, t.description,
            {"new": "Новая", "on_moderation": "На модерации", "in_progress": "В работе", "done": "Завершено", "cancelled": "Отменено"}.get(t.status, t.status),
            {"low": "Низкий", "medium": "Средний", "high": "Высокий"}.get(t.priority, t.priority),
            t.deadline.strftime("%d.%m.%Y %H:%M") if t.deadline else "",
            users_map.get(t.author_id, ""),
            users_map.get(t.assignee_id, ""),
            cats_map.get(t.category_id, ""),
            t.created_at.strftime("%d.%m.%Y %H:%M") if t.created_at else "",
            t.updated_at.strftime("%d.%m.%Y %H:%M") if t.updated_at else "",
            comments_text,
        ])
    csv_content = output.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=tickets.csv"},
    )


@router.get("/export/xlsx")
def export_xlsx(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    category_id: Optional[int] = None,
    assignee_id: Optional[int] = None,
    author_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(get_optional_user),
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    svc = TicketService(db)
    comment_svc = CommentService(db)
    items_raw, _ = svc.get_all_filtered(skip=0, limit=10000, status=status, priority=priority, category_id=category_id, assignee_id=assignee_id, author_id=author_id)

    users_map = {u.id: u.name for u in db.query(User).all()}
    cats_map = {c.id: c.name for c in db.query(Category).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["ID", "Название", "Описание", "Статус", "Приоритет", "Дедлайн", "Автор", "Исполнитель", "Категория", "Создана", "Обновлена", "Комментарии"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, t in enumerate(items_raw, 2):
        comments = comment_svc.get_by_ticket(t.id)
        comments_text = " | ".join(
            f"{c.author_name} ({c.created_at.strftime('%d.%m.%Y')}): {c.text}"
            for c in comments
        ) if comments else ""
        ws.cell(row=i, column=1, value=t.id).border = thin_border
        ws.cell(row=i, column=2, value=t.title).border = thin_border
        ws.cell(row=i, column=3, value=t.description).border = thin_border
        ws.cell(row=i, column=4, value={"new": "Новая", "on_moderation": "На модерации", "in_progress": "В работе", "done": "Завершено", "cancelled": "Отменено"}.get(t.status, t.status)).border = thin_border
        ws.cell(row=i, column=5, value={"low": "Низкий", "medium": "Средний", "high": "Высокий"}.get(t.priority, t.priority)).border = thin_border
        ws.cell(row=i, column=6, value=t.deadline.strftime("%d.%m.%Y %H:%M") if t.deadline else "").border = thin_border
        ws.cell(row=i, column=7, value=users_map.get(t.author_id, "")).border = thin_border
        ws.cell(row=i, column=8, value=users_map.get(t.assignee_id, "")).border = thin_border
        ws.cell(row=i, column=9, value=cats_map.get(t.category_id, "")).border = thin_border
        ws.cell(row=i, column=10, value=t.created_at.strftime("%d.%m.%Y %H:%M") if t.created_at else "").border = thin_border
        ws.cell(row=i, column=11, value=t.updated_at.strftime("%d.%m.%Y %H:%M") if t.updated_at else "").border = thin_border
        ws.cell(row=i, column=12, value=comments_text).border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    xlsx_bytes = io.BytesIO()
    wb.save(xlsx_bytes)
    xlsx_bytes.seek(0)

    return StreamingResponse(
        iter([xlsx_bytes.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tickets.xlsx"},
    )
